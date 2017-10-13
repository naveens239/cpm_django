from django.shortcuts import render, render_to_response,redirect
from django.http import HttpResponseRedirect, HttpResponse
from django.contrib.auth.models import User
from django.core.urlresolvers import reverse
from django.core import serializers
from .forms import ContactForm, SignUpForm,AddNewProjectForm,ProjectStageForm,TeamAddForm,TeamEditForm, OrderListingForm, TrackingForm
from .forms import ProjectPlanForm,ScheduleForm,ScheduleEditForm, MaterialForm, PrototypeForm, ScheduleCommentForm, VendorListingForm,MaterialCommentForm, ProcessURLForm,ExcelForm
from django.core.mail import send_mail
from django.conf import settings
from .models import Project, Stage, StageSetting, Team, Role,Plan, Schedule, Material, OrderStatus, Prototype, ScheduleComment, MaterialComment
from .models import TrackingInfo, Courier, CategoryList, VendorList, OrderPriority,ReadCommentTrack
from product_details_scraper import ReadAsin
import json, ast, datetime,re,mimetypes,os
import openpyxl
import aftership as aftership

api = aftership.APIv4("8a4aab03-8132-4fae-aca8-48c054964e14")
def excel_courier_loader():
    wb = openpyxl.load_workbook("C:\Users\acer-pc\Documents\ET\my_project_folder\couriers.xlsx")
    sheet = wb.get_sheet_by_name('Sheet1')
    for row in range(2, sheet.max_row + 1):
       name = sheet['A'+str(row)].value
       slug = sheet['B'+str(row)].value
       try:
          p = Courier(name=name,slug=slug)
          p.save() 
       except Exception as e:
        print str(e)
# Create your views here.
def home(request):
    title = "Sign Up Now"
    # if request.user.is_authenticated():
         # title = "Welcome %s" %(request.user)
    form = SignUpForm(request.POST or None)
    context = {
        "title": title,
        "form" : form
    }
   
    if form.is_valid():
        form.save()
        context = {
           "title": "Thank you"
        }
    if request.user.is_authenticated() and request.user.is_staff:
        context={
            "queryset":[123,456]
        }    
    return render(request, "home.html",context)


def contact(request):
    title = 'Contact Us'
    title_align_center  = True
    form = ContactForm(request.POST or None)
    if form.is_valid():
        # for key in form.cleaned_data:
        #     print key
        #     print form.cleaned_data.get(key)
        form_email = form.cleaned_data.get("email")
        form_message = form.cleaned_data.get("message")
        form_first_name = form.cleaned_data.get("first_name")
        #print form.cleaned_data
        subject = 'Site contact form'
        from_email = settings. EMAIL_HOST_USER
        to_email = [from_email,'nbengt14@gmail.com']
        email_message =  "%s : %s via %s"%(
            form_first_name,
            form_message,
            form_email)
        send_mail(subject,
         email_message,
         from_email,
         to_email,
         fail_silently= False)
    context = {
        "form" : form,
        "title": title,
        "title_align_center":title_align_center,
    }
    return render(request, "forms.html", context)

# Create your views here.
def about(request):


    return render(request, "about.html",{})

def profile(request):
    project_data = Project.objects.all()
    order_listing_form = OrderListingForm(request.POST or None)
    listing_message=""
    
    if request.method=='POST' and 'register_listing' in request.POST:
      print 'in order listing'
      print request.POST
      try:
          if order_listing_form.is_valid():
              print 'is valid'
              category_choice =  order_listing_form.cleaned_data.get("category_choice")
              if category_choice == "Others":
                order_category = order_listing_form.cleaned_data.get("category")
              else:
                order_category = category_choice
              sub_category_choice =  order_listing_form.cleaned_data.get("sub_category_choice")
              if sub_category_choice == "Others":
                order_sub_category = order_listing_form.cleaned_data.get("sub_category")
              else:
                order_sub_category = sub_category_choice

             #print 'SUb is.....', order_sub_category

              if (order_category != None and order_category !="" and order_category != "Others") and (order_sub_category != None and order_sub_category != "" and order_sub_category != "Others"):
                #print 'order category is', order_category
                #print 'orer sub is ', order_sub_category
                category_data = CategoryList.objects.filter(category=order_category, sub_category=order_sub_category)
                #print 'data', category_data
                if not category_data:
                    #print 'new cat data'
                    category_data = CategoryList(category=order_category, sub_category= order_sub_category)
                    category_data.save()
                    listing_message= "Category Saved Successfully"
                    category_data = CategoryList()
                elif not CategoryList.objects.filter(category=order_category, sub_category=order_sub_category).exists():
                    #print 'cat data 2'
                    category_data = CategoryList(category=order_category, sub_category= order_sub_category)
                    category_data.save()
                    category_data = CategoryList()
                    listing_message= "Category Saved Successfully"
                else:
                    #print 'cat condition 3'
                    listing_message= "Category pair already exists"
             
              else:
                 #print 'errrrors'
                 #print 'nothing',order_listing_form.errors
                 listing_message = "Category/ Sub category can't be empty.Please provide valid entry"
          else:
            if (request.POST.get('category_choice')=="Others" and request.POST.get('category') == "") or (request.POST.get('sub_category_choice')=="Others" and request.POST.get('sub_category') == ""):
               listing_message = "Category/ Sub category can't be empty.Please provide valid entry"
            print order_listing_form.errors
      except Exception as e:
          print str(e)  
    order_listing_form = OrderListingForm()
    context={
            "form":project_data,
            "order_listing_form":order_listing_form,
            "listing_message":listing_message,
        }

    return render(request, "profile.html",context)    

def get_vendors(request):
    print 'in here sub category filtering'
    ven_data = VendorList.objects.all().values_list("vendor_name","vendor_name").distinct()
    print 'ven dat is ', ven_data
    cat=[]
    for sc in ven_data:
       cat.append(sc[0])
    print cat
    return HttpResponse(json.dumps(cat), content_type="application/json")

def get_category(request):
    print 'in here sub category filtering'
    cat_data = CategoryList.objects.all().values_list("category","category").distinct()
    print 'cat dat is ', cat_data
    cat=[]
    for sc in cat_data:
       cat.append(sc[0])
    print cat
    return HttpResponse(json.dumps(cat), content_type="application/json")
def get_sub_category(request,category):
    print 'in here sub category filtering',category
    cat_data = CategoryList.objects.filter(category=category)
    print 'cat dat is ', cat_data
    sub_cat=[]
    for sc in cat_data:
       sub_cat.append(sc.sub_category)
    print sub_cat
    #form = OrderListingForm(request.POST)
    #form.fields['sub_category_choice'].choices = [(sub_cat,sub_cat)]
    return HttpResponse(json.dumps(sub_cat), content_type="application/json")
    #return HttpResponse(request,form)
def register_vendor(request):
    vendor_listing_form = VendorListingForm(request.POST or None)
    message = ""
    if request.method=='POST' and 'vendor_listing' in request.POST:
      print 'in vendor', request.POST
      if vendor_listing_form.is_valid():
         vendor_name = vendor_listing_form.cleaned_data.get("vendor_name")
         GSTIN = vendor_listing_form.cleaned_data.get("GSTIN")
         address = vendor_listing_form.cleaned_data.get("address")
         contact_person = vendor_listing_form.cleaned_data.get("contact_person")
         contact_num = vendor_listing_form.cleaned_data.get("contact_num")
         website = vendor_listing_form.cleaned_data.get("website")
         vendor_data = VendorList(vendor_name=vendor_name,address=address,GSTIN=GSTIN,contact_person=contact_person,contact_num=contact_num,website=website)
         vendor_data.save()
      else:
         print vendor_listing_form.errors
    vendor_data = VendorList.objects.all()
    

    if request.method=='POST' and 'vendor_delete' in request.POST:
       id_ = request.POST.get('model_instance')
       print'Name issss',id_
       vendor_data = VendorList.objects.get(id=id_)
       vendor_data.delete()
       message = "Vendor Deleted Successfully"
    
    if request.method=='POST' and 'vendor_edit' in request.POST:
      vendor_edit_form = VendorListingForm(request.POST or None)
      if vendor_edit_form.is_valid():   
            print request.POST
            vendor_name = vendor_edit_form.cleaned_data.get('vendor_name')
            address = vendor_edit_form.cleaned_data.get('address')
            GSTIN = vendor_edit_form.cleaned_data.get('GSTIN')
            contact_person = vendor_edit_form.cleaned_data.get('contact_person')
            contact_num = vendor_edit_form.cleaned_data.get('contact_num')
            website = vendor_edit_form.cleaned_data.get('website')
            id_ = request.POST.get("model_instance")
            print 'before try !!!!!!!', id_
            try:
              p = VendorList.objects.get(id=id_)
              if vendor_name:
                  p.vendor_name = vendor_name
              if address:
                  p.address = address
              if GSTIN:
                  p.GSTIN = GSTIN
              if contact_person:
                  p.contact_person = contact_person
              if contact_num:
                  p.contact_num = contact_num
              if website:
                  p.website = website

              p.save()
              message="Vendor details Updated Successfully"
            except VendorList.DoesNotExist:
               print 'in except'
               p = VendorList(vendor_name=vendor_name,address=address,GSTIN=GSTIN,contact_num=contact_num,contact_person=contact_person,website=website)
               p.save()
               message="Vendor details Updated Successfully"
      else:
            print 'errror in herereeeee'
            print request.method
            message="Vendor Update Not Successful"
            print 'something went wrong !!!!!!!!'   

    context={
      'vendor_data':vendor_data,
      'vendor_listing_form':vendor_listing_form,
      'vendor_message': message,
    }

    return render(request, "register_vendor.html",context)  
  
def addproject(request):
    #print request.POST
    title = 'Add New Project'
    title_align_center  = True
    form = AddNewProjectForm(request.POST or None)
    message=""
    if form.is_valid():
        name = form.cleaned_data.get("name")
        start_date = form.cleaned_data.get("start_date")
        end_date = form.cleaned_data.get("end_date")
        # for key in form.cleaned_data:
        #     print key
        #     print form.cleaned_data.get(key)
        # form_prj_name = form.cleaned_data.get("name")
        # form_prj_status = form.cleaned_data.get("status")
        # form_prj_completion = form.cleaned_data.get("completion")
        # form.save()
        project_data = Project.objects.all()
      
        p = Project(name=name,status="Ongoing",completion = 0,start_date=start_date,end_date=end_date)
        if start_date > end_date:
            message = "Project End date cannot be earlier than Start date"
        else:
            message = "Project Added Successfully"
            p.save()
        context={
            "form":project_data,
            "message":message,               
        }
        return redirect('/profile/',context) 
        print'form data', form
    context = {
        "form" : form,
        "title": title,
        "title_align_center":title_align_center,
    }
    return render(request, "forms.html", context)

def netproject(request):
    print request.POST
    try:
         material_data = None
         material_data = Material.objects.all()
    except:
         material_data = None
    context={
        "material_data":material_data,
    }
    
    return render(request,"net_project_view.html",context)


def projectoverview(request,name):
    print request.POST
    print name
    project_data = Project.objects.get(name=name)
    try:
      team_data = None
      team_data = Team.objects.filter(project_name=project_data.id)
    except Team.DoesNotExist:
       team_data = None
    context={
        "project_data":project_data,
        "team_data":team_data,
    }
    
    return render(request,"project_overview.html",context)

def projectsettings(request,name):
    
   if '/' in name:
        name, rest = name.split("/")
   project_data = Project.objects.get(name=name)
   stage_data = Stage.objects.all()
   role_data = Role.objects.all()
   print 'hereeee', request.POST
   #print 'get data',request.GET
   if request.method=='POST' and 'stage' in request.POST:
        form_stage = ProjectStageForm(request.POST or None)
        if form_stage.is_valid():
            print 'form is valid - ^^^^^^^^^^^^^^^^^^'
            list = request.POST.getlist('stage_item')
            list = ast.literal_eval(json.dumps(list))
            check_list  = list
            checked_items=check_list
            print 'checked_items',checked_items
            project_completion = int(round((float(len(checked_items))/len(stage_data))*100.0))
            if project_completion == 100:
                project_status = "Completed"
            else:
                project_status = "Ongoing"
            try:
                print'in herere'
                p = StageSetting.objects.get(project_name=project_data.id)
                print 'p is ',p.checked_items
                p.checked_items = checked_items
                p.total_true_checks = len(checked_items)
                project_data.completion=project_completion
                project_data.status = project_status
                project_data.save()
                p.save()    
            except StageSetting.DoesNotExist:
                print 'in hrerer except'
                p = StageSetting(project_name=project_data,total_true_checks=len(checked_items),checked_items=checked_items)
                project_data.completion=project_completion
                project_data.status = project_status
                project_data.save()
                p.save()    
        else:
            print request.method
            print form_stage.errors
            print 'something went wrong !!!!!!!!'   
 
   if request.method=='POST' and 'save' in request.POST:
         team_add_form = TeamAddForm(request.POST or None)
         if team_add_form.is_valid():
            print 'teammmmm', request.POST
            role_id = team_add_form.cleaned_data.get("role_name")
            member_name = team_add_form.cleaned_data.get("member_name")
            print'in herere'
            role= Role.objects.get(id=role_id)
            print 'role is ', role.name
            p = Team(project_name=project_data,member_name=member_name,member_role=role)
            p.save() 
                # p = Team.objects.get(project_name=project_data.id)
                # p.member_name = member_name
                # p.role_name = role_name
                # p.project_name = 
                # p.save()    
   if request.method=='POST' and 'edit' in request.POST:
          team_edit_form = TeamEditForm(request.POST or None)
          
          if team_edit_form.is_valid():
            try:
               # list = request.POST.getlist('stage_item')
               # list = ast.literal_eval(json.dumps(list))
                id_ = team_edit_form.cleaned_data.get("model_instance")
                #team = Team.objects.get(pk=id)
                print'Name issss',id_
                item_data = Team.objects.get(id=id_)
                item_data.member_name = team_edit_form.cleaned_data.get("member_name")
                item_data.member_role = Role.objects.get(id=team_edit_form.cleaned_data.get("role_name"))
                item_data.save()
            except Team.DoesNotExist:
                print 'in hrerer except'
                p = Team(project_name=project_data,member_name=member_name,role_name=member_role)
                p.save()    

   if request.method=='POST' and 'delete' in request.POST:
                id_ = request.POST.get('model_instance')

                print'Name issss',id_
                item_data = Team.objects.get(id=id_)
                item_data.delete()
  

   form_stage = ProjectStageForm(request.POST or None)
   team_add_form = TeamAddForm(request.POST or None)
   team_edit_form = TeamEditForm(request.POST or None)
   try:
       team_data = None
       team_data = Team.objects.filter(project_name=project_data.id)
   except Team.DoesNotExist:
       team_data = None
   try:
       project_stage_data = None
       project_stage_data = StageSetting.objects.get(project_name=project_data.id)      
       form_stage = ProjectStageForm(initial={'stage_item': project_stage_data.checked_items})
   except StageSetting.DoesNotExist:
       project_stage_data = None

    
   context={
        "stage_form":form_stage,
        "team_add_form":team_add_form,
        "team_edit_form":team_edit_form,
        "project_data":project_data,
        "stage_data":stage_data,
        "project_stage_data":project_stage_data,
        "team_data":team_data,
        "role_data":role_data,
   }
    
   return render(request,"project_settings.html",context)

def projectdetails(request,name):
      if '/' in name:
            name, rest = name.split("/")
      print request.POST
      project_data = Project.objects.get(name=name)
      role_data = Role.objects.all()
      order_status_data = OrderStatus.objects.all()
      task_message = ""
      plan_message = ""
      order_message = ""
      prototype_message =""
      process_message = ""
      status_message = ""

      if request.method =='POST' and 'status' in request.POST:
        print 'status data pass success',request
        order_id = request.POST.get("order_id")
        status_val = request.POST.get("status")
        m = Material.objects.get(id=order_id)
        old_status = m.order_status.status_id
        if old_status < 500:
           new_status = old_status+100 # change status
        elif old_status == 500 and status_val =="working":
           new_status = 600
        elif old_status == 500 and status_val =="not working":
           new_status = 700
        m.order_status = OrderStatus.objects.get(status_id=new_status)
        m.save(update_fields=['order_status'])
        status_message="Status updated Successfully"    
        o = OrderStatus.objects.get(status_id=old_status)
        n = OrderStatus.objects.get(status_id=new_status)
        author = request.user
        comment = "Changed order status from \""+o.name+"\" to \""+n.name+"\""
        c = MaterialComment(material=m,comment=comment,author=author)
        c.save()
        status_message="Status updated successfully"
        request.session['order_message'] = status_message
        return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))
      if request.method=='POST' and 'plan' in request.POST:
        plan_form = ProjectPlanForm(request.POST or None)
        if plan_form.is_valid():
            project_plan = plan_form.cleaned_data.get("project_plan")
            business_plan = plan_form.cleaned_data.get("business_plan")
            wiki_link = plan_form.cleaned_data.get("wiki_link")
            prototype_url = plan_form.cleaned_data.get("prototype_url")
            weblink_url = plan_form.cleaned_data.get("weblink_url")
            try:               
                p = Plan.objects.get(project_name=project_data.id)
                p.project_plan = project_plan
                p.business_plan = business_plan
                p.wiki_link = wiki_link
                p.prototype_url = prototype_url
                p.weblink_url = weblink_url
                p.save()    
                if not project_plan and not business_plan and not wiki_link:
                  p.delete()
                plan_message="Plan Updated Successfully"
                request.session['plan_message'] = plan_message
                return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))
            except Plan.DoesNotExist:
                p = Plan(project_name=project_data,project_plan=project_plan,
                  business_plan=business_plan,wiki_link=wiki_link,prototype_url=prototype_url,weblink_url=weblink_url)
                p.save()
                plan_message="Plan Added Successfully"
                request.session['plan_message'] = plan_message
                return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))
      if request.method=='POST' and 'schedule_save' in request.POST:
         schedule_form = ScheduleForm(request.POST or None)
         print request.POST
         print 'hererere in task 11111'

         if schedule_form.is_valid():   
            print 'hererere in task 2222'
            task_name = schedule_form.cleaned_data.get('task_name')
            assigned_id = schedule_form.cleaned_data.get('assigned_to')
            start_date = schedule_form.cleaned_data.get('start_date')
            end_date = schedule_form.cleaned_data.get('end_date')
            role_assigned_to = Role.objects.get(id=assigned_id)     
            p = Schedule(project_name=project_data,assigned_to=role_assigned_to,start_date=start_date,end_date=end_date,task_name=task_name)
            if p.end_date < p.start_date:
               task_message="Task End date cannot be earlier than Start date"   
            else:
               p.save()
               task_message="Task Added Successfully"
            request.session['task_message'] = task_message
            return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))
         else:
            print 'errror in herereeeee'
            print request.method
            task_message="Task could not be added"
            print schedule_form.errors
            request.session['task_message'] = task_message
            return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))
      if request.method=='POST' and 'schedule_edit' in request.POST:
         schedule_edit_form = ScheduleEditForm(request.POST or None)
         print request.POST

         if schedule_edit_form.is_valid():   

            task_name = schedule_edit_form.cleaned_data.get('task_name')
            assigned_id = schedule_edit_form.cleaned_data.get('assigned_to')
            start_date = schedule_edit_form.cleaned_data.get('start_date')
            end_date = schedule_edit_form.cleaned_data.get('end_date')
            id_ = schedule_edit_form.cleaned_data.get("model_instance")
            print 'before try !!!!!!!'
            try:
               p = Schedule.objects.get(id=id_)
               if task_name:
                  p.task_name = task_name
               if assigned_id:
                  p.assigned_to = Role.objects.get(id=assigned_id)
               if start_date:
                  p.start_date = start_date
               else:
                  p.start_date = p.start_date.date()   
               if end_date:
                  p.end_date = end_date
               else:
                  p.end_date = p.end_date.date()   
               if p.end_date < p.start_date:
                  task_message="Task End date cannot be earlier than Start date"   
               else:
                  p.save()
                  task_message="Task Updated Successfully"
               request.session['task_message'] = task_message
               return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))
            except Schedule.DoesNotExist:
               role_assigned_to = Role.objects.get(id=assigned_id)     
               p = Schedule(project_name=project_data,assigned_to=role_assigned_to,start_date=start_date,end_date=end_date,task_name=task_name)
               p.save()
               task_message="Task Updated Successfully"
               request.session['task_message'] = task_message
               return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))
         else:
            print 'errror in herereeeee'
            print request.method
            task_message="Task Update Not Successful"
            print 'something went wrong !!!!!!!!'   
            request.session['task_message'] = task_message
            return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))
      if request.method=='POST' and 'schedule_comment' in request.POST:
         schedule_comment_form = ScheduleCommentForm(request.POST or None)
         print request.POST

         if schedule_comment_form.is_valid():   

            author = schedule_comment_form.cleaned_data.get('author')
            print 'author is',author
            comment = request.POST.get('comment_detail')
            print 'comment is',comment
            id_ = schedule_comment_form.cleaned_data.get("model_instance")

            schedule = Schedule.objects.get(id=id_)     
            p = ScheduleComment(schedule=schedule,comment=comment,author=author)
            if comment :
                p.save()
            else:
               task_message="Comment could not be added"
            request.session['task_message'] = task_message
            return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))
         else:
            print 'errror in herereeeee'
            print request.method
            task_message="Comment could not be added"
            print schedule_comment_form.errors
            request.session['task_message'] = task_message
            return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))

      if request.method=='POST' and 'order_comment' in request.POST:
         material_comment_form = MaterialCommentForm(request.POST or None)
         print request.POST

         if material_comment_form.is_valid():   

            author = material_comment_form.cleaned_data.get('author')
            print 'author is',author
            comment = request.POST.get('comment_detail')
            print 'comment is',comment
            id_ = material_comment_form.cleaned_data.get("model_instance")

            material = Material.objects.get(id=id_)     
            p = MaterialComment(material=material,comment=comment,author=author)
            if comment :
                p.save()
            else:
               order_message="Comment could not be added"
            return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))
         else:
            print 'errror in herereeeee'
            print request.method
            order_message="Comment could not be added"
            print material_comment_form.errors
            return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))

      if request.method=='POST' and 'schedule_delete' in request.POST:
                id_ = request.POST.get('model_instance')
                print'Name issss',id_
                schedule_data = Schedule.objects.get(id=id_)
                schedule_data.delete()
                task_message="Task Deleted Successfully"
                return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))
      
      if request.method =='POST' and 'download_excel' in request.POST:
          file_full_path = "cpm/orders.xlsx"
          with open(file_full_path,'rb') as f:
              data = f.read()
          response = HttpResponse(data, content_type=mimetypes.guess_type(file_full_path)[0])
          response['Content-Disposition'] = "attachment; filename=orders"
          response['Content-Length'] = os.path.getsize(file_full_path)
          return response
      if request.method=='POST' and 'excel_data' in request.POST:
        print 'hereeeeee'
        excel_form = ExcelForm(request.POST, request.FILES)
         
         
        print request.POST
        if excel_form.is_valid() and request.POST.get('excel_file')!="":
          print 'vaaalidddd'
          excel = request.FILES['excel_file']
          print excel
          name = excel.name
          extension=[]
          extension = name.split('.')
          print extension
          if extension[1] in ('xlsx','xls'):
            print "right extension"
            wb = openpyxl.load_workbook(excel)
            sheet = wb.get_sheet_by_name('Sheet1')
            if sheet['A'+str(1)].value == "Id":
                for row in range(2, sheet.max_row + 1):
                  id_ = sheet['A'+str(row)].value
                  category = sheet['C'+str(row)].value
                  subcategory = category.split('-')[1]
                  category = category.split('-')[0]
                  item = sheet['D'+str(row)].value
                  vendor = sheet['E'+str(row)].value
                  url = sheet['F'+str(row)].value
                  quantity = sheet['G'+str(row)].value
                  currprice = sheet['H'+str(row)].value
                  currency = currprice[0:2]
                  price = currprice[2:]
                  hsn = sheet['I'+str(row)].value
                  print category, subcategory, item, vendor, url, quantity, currprice, currency, price, hsn
                  update_list =[]
                  comment=""
                  try:
                      p = Material.objects.get(id=id_)

                      if str(category) != p.order_category:
                        comment=comment+" Changed category from \""+p.order_category+"\" to\""+category+"\"."
                        p.order_category = category
                        update_list.append('order_category')
                      if str(subcategory) != p.order_sub_category:
                        comment=comment+" Changed subcategory from \""+p.order_sub_category+"\" to\""+subcategory+"\"."
                        p.order_sub_category = subcategory
                        update_list.append('order_sub_category')
                      if str(item) != p.order_item:
                        comment=comment+" Changed item from \""+p.order_item+"\" to\""+item+"\"."
                        p.order_item = item
                        update_list.append('order_item')
                      if str(url) != p.order_item_url:
                        comment=comment+" Changed URL from \""+str(p.order_item_url)+"\" to\""+str(url)+"\"."
                        p.order_item_url = url
                        update_list.append('order_item_url')
                      if str(vendor) != p.order_vendor:
                        comment=comment+" Changed vendor from \""+p.order_vendor+"\" to\""+vendor+"\"."
                        p.order_vendor = vendor
                        update_list.append('order_vendor')
                      if int(quantity) != p.order_quantity:
                        comment=comment+" Changed quantity from \""+str(p.order_quantity)+"\" to\""+str(quantity)+"\"."
                        p.order_quantity = quantity
                        update_list.append('order_quantity')
                      # if currency !=p.order_currency:
                      #   p.order_currency = currency
                      if float(price) != p.order_unit_price:
                        print type(price), type(p.order_unit_price)
                        comment=comment+" Changed price from \""+str(p.order_unit_price)+"\" to\""+str(price)+"\"."
                        p.order_unit_price = price
                        if p.order_unit_price == "":
                            p.order_unit_price =0.00
                        update_list.append('order_unit_price')
   
                      
                      if str(hsn) != p.order_hsn:
                        comment=comment+" Changed HSN code from \""+str(p.order_hsn)+"\" to\""+str(hsn)+"\"."
                        p.order_hsn = hsn
                        update_list.append('order_hsn')

                      p.save(update_fields=update_list)
                      if comment !="":
                        print 'comments are not none'
                        author = request.user
                        c = MaterialComment(material=p,comment=comment,author=author)
                        c.save()
                        all_users = User.objects.all()
                        for i in all_users:
                          print i.username
                          t = ReadCommentTrack(order_id=p.id,project_id = project_data.id,user_name=i.username,read_flag="N")
                          t.save()
                      try:
                        check_vendor = VendorList.objects.get(vendor_name=vendor)
                        print 'vendor',check_vendor
                        if check_vendor=="":
                            v = VendorList(vendor_name=vendor,address="",GSTIN="",contact_person="",contact_num=9999999999,website="")
                            v.save()
                      except VendorList.DoesNotExist:
                          v = VendorList(vendor_name=vendor,address="",GSTIN="",contact_person="",contact_num=9999999999,website="")
                          v.save()
                      try:
                          check_sub_category = CategoryList.objects.get(sub_category=subcategory)
                          print 'category',check_sub_category
                          if check_sub_category=="":
                              c = CategoryList(category=category,sub_category=subcategory)
                              c.save()
                      except CategoryList.DoesNotExist:
                              c = CategoryList(category=category,sub_category=subcategory)
                              c.save()
                      order_message="Excel Order Updated Successfully"
                      
                  except Material.DoesNotExist:
                      print 'in exception blck'
                      status = OrderStatus.objects.get(status_id = 100)
                      priority = OrderPriority.objects.get(priority_id=200)
                      elt = "1d"
                      p = Material(project_name=project_data,order_category=category,\
                                order_sub_category=subcategory,order_item=item,\
                                order_item_url=url,order_vendor = vendor,order_quantity=int(quantity),\
                                order_currency = currency,order_unit_price=float(price), order_status=status, \
                                est_lead_time=elt,order_priority=priority, order_hsn=hsn,author=request.user)
                      p.save()
                      order_message="Excel Order Updated Successfully"
                  print order_message
                  
                request.session['order_message'] = order_message
                return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))

            else:
                for row in range(2, sheet.max_row + 1):
                  category = sheet['A'+str(row)].value
                  subcategory = sheet['B'+str(row)].value
                  item = sheet['C'+str(row)].value
                  vendor = sheet['D'+str(row)].value
                  url = sheet['E'+str(row)].value
                  quantity = sheet['F'+str(row)].value
                  price = sheet['G'+str(row)].value
                  status = OrderStatus.objects.get(status_id=100)
                  priority = OrderPriority.objects.get(priority_id=200)
                  currency = "" #empty
                  author = request.user
                  elt = "1d"
                  hsn = sheet['H'+str(row)].value
                  if price =="":
                      price =0.00
                  if currency =="":
                      currency="Rs"
                  print category, subcategory, item, vendor, url, quantity, price, status, currency, hsn
                  try:
                      check_vendor = VendorList.objects.get(vendor_name=vendor)
                      print 'vendor',check_vendor
                      if check_vendor=="":
                          v = VendorList(vendor_name=vendor,address="",GSTIN="",contact_person="",contact_num=9999999999,website="")
                          v.save()
                  except VendorList.DoesNotExist:
                      v = VendorList(vendor_name=vendor,address="",GSTIN="",contact_person="",contact_num=9999999999,website="")
                      v.save()
                  try:
                      check_sub_category = CategoryList.objects.get(sub_category=subcategory)
                      print 'category',check_sub_category
                      if check_sub_category=="":
                          c = CategoryList(category=category,sub_category=subcategory)
                          c.save()
                  except CategoryList.DoesNotExist:
                          c = CategoryList(category=category,sub_category=subcategory)
                          c.save()
                  try:
                      p = Material(project_name=project_data,order_category=category,\
                                    order_sub_category=subcategory,order_item=item,order_vendor=vendor,\
                                    order_item_url=url,order_quantity=quantity,est_lead_time=elt,\
                                    order_currency=currency,order_unit_price=price, order_status=status, order_hsn=hsn,order_priority=priority,author=author)
                      p.save() 
                      order_message =  "Excel sheet successfully processed."
                      request.session['order_message'] = order_message
                      return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name]))
                  except Exception as e:
                    print str(e)
                    order_message =  "Order number "+ str(row)+" could not be saved."
        else:
           print 'form INVALID'
           order_message = "Please upload .xls or .xlsx file only"
           request.session['order_message'] = order_message
           print excel_form.errors
           return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name]))


      if request.method=='POST' and 'process_url' in request.POST:
         process_url_form = ProcessURLForm(request.POST or None)
         extracted_data=[]
         print request.POST
         if process_url_form.is_valid():
            url = process_url_form.cleaned_data.get('process_link')
            extracted_data = ReadAsin(url)
            print extracted_data
            if len(extracted_data)==0:
              order_message ="Oops there has been an error. Please add order manually"
              print 'Please add order manually'
            else:
              print 'Found dataaaaaaaaaaaaa'
              try:
                  order_category = extracted_data[0]['CATEGORY']
                  if order_category == "":
                    order_category = "Unknown"
                  order_sub_category = extracted_data[0]['SUBCATEGORY']
                  if order_sub_category == "":
                    order_sub_category = "Unknown"
                  order_item = extracted_data[0]['ITEM']
                  if order_item == "":
                    order_item = "Nil"
                  order_vendor = extracted_data[0]['VENDOR']
                  order_item_url = extracted_data[0]['URL']
                  order_quantity = 1
                  order_unit_price = extracted_data[0]['ORIGINAL_PRICE']
                  if order_unit_price =="":
                    order_unit_price=0.00
                  author = request.user
                  est_lead_days = "1d"

                  # print 'string is',price_string
                  # match = re.search(r'([\D]+)([\d,.]+)', price_string)
                  # output = (match.group(1), match.group(2).replace(',',''))
                  # print 'output is',output[1]
                  # order_unit_price = output[1]
                  order_currency = extracted_data[0]['CURRENCY']
                  if order_currency=="":
                    order_currency='Rs'
                 
                  order_priority = OrderPriority.objects.get(priority_id=300)
                  order_status = OrderStatus.objects.get(status_id=100)
                  try:
                     check_vendor = VendorList.objects.get(vendor_name=order_vendor)
                     print 'vendor',check_vendor
                     if check_vendor=="":
                        v = VendorList(vendor_name=order_vendor,address="",GSTIN="",contact_person="",contact_num=9999999999,website="")
                        v.save()
                  except VendorList.DoesNotExist:
                     v = VendorList(vendor_name=order_vendor,address="",GSTIN="",contact_person="",contact_num=9999999999,website="")
                     v.save()
                  try:
                     check_sub_category = CategoryList.objects.get(sub_category=order_sub_category)
                     print 'category',check_sub_category
                     if check_sub_category=="":
                        c = CategoryList(category=order_category,sub_category=order_sub_category)
                        c.save()
                  except CategoryList.DoesNotExist:
                        c = CategoryList(category=order_category,sub_category=order_sub_category)
                        c.save()
                  p = Material(project_name=project_data,order_category=order_category,author = author,est_lead_time=est_lead_days,\
                                order_sub_category=order_sub_category,order_item=order_item,order_vendor=order_vendor,\
                                order_item_url=order_item_url,order_quantity=order_quantity,\
                                order_currency=order_currency,order_unit_price=order_unit_price, order_status=order_status,order_priority=order_priority,order_hsn="")
                  p.save(),
                  order_message="Order Placed Successfully"
                  request.session['order_message'] = order_message
                  return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))
              except Exception as e:
                print str(e)
                order_message="Oops there has been an error. Please add order manually"
                request.session['order_message'] = order_message
                return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))
               
      if request.method=='POST' and 'material_save' in request.POST:
            material_form = MaterialForm(request.POST or None)
            print request.POST
            print 'hererere in order 11111'
            if material_form.is_valid():   
                print 'hererere in order 2222'
                order_category = material_form.cleaned_data.get('order_category')
                order_sub_category = material_form.cleaned_data.get('order_sub_category')
                order_item = material_form.cleaned_data.get('order_item')
                order_vendor = material_form.cleaned_data.get('order_vendor')
                order_item_url = material_form.cleaned_data.get('order_item_url')
                order_quantity = material_form.cleaned_data.get('order_quantity')
                order_currency = material_form.cleaned_data.get('order_currency')
                order_hsn = material_form.cleaned_data.get('order_hsn')
                if order_currency=="":
                  order_currency='Rs'
                order_unit_price = material_form.cleaned_data.get('order_unit_price')
                if order_unit_price =="":
                  order_unit_price=0.00
                order_status = OrderStatus.objects.get(status_id=100)
                priority_id = material_form.cleaned_data.get('order_priority')
                order_priority = OrderPriority.objects.get(priority_id=priority_id)
                author = request.user
                est_lead_time = str(material_form.cleaned_data.get('est_lead_num'))+material_form.cleaned_data.get('est_lead_days')
                p = Material(project_name=project_data,order_category=order_category,\
                              order_sub_category=order_sub_category,order_item=order_item,order_vendor=order_vendor,\
                              order_item_url=order_item_url,order_quantity=order_quantity,\
                              order_currency=order_currency,order_unit_price=order_unit_price,\
                              order_status=order_status,author=author,est_lead_time=est_lead_time,order_priority=order_priority,order_hsn=order_hsn)
                p.save()
                if order_message=="":
                    order_message="Order Placed Successfully"
            else:
                print 'errror in herereeeee'
                print request.method
                if order_message=="":
                  order_message="Order could not be placed"
                print material_form.errors
            request.session['order_message'] = order_message
            return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))
      if request.method=='POST' and 'order_edit' in request.POST:
         material_form = MaterialForm(request.POST or None)
         print request.POST

         if material_form.is_valid():   

            item = material_form.cleaned_data.get('order_item')
            print 'item is', item 
            category = material_form.cleaned_data.get('order_category')
            sub_category = material_form.cleaned_data.get('order_sub_category')
            quantity = material_form.cleaned_data.get('order_quantity')
            currency = material_form.cleaned_data.get('order_currency')
            price = material_form.cleaned_data.get('order_unit_price')
            vendor = material_form.cleaned_data.get('order_vendor')
            url = material_form.cleaned_data.get('order_item_url')
            id_ = request.POST.get("model_instance")
            status_id = request.POST.get('order_status')
            est_lead_num = request.POST.get('est_lead_num')
            est_lead_days = request.POST.get('est_lead_days')
            priority = material_form.cleaned_data.get('order_priority')
            hsn = material_form.cleaned_data.get('order_hsn')
            update_list =[]
            comment=""
            print 'status_id is ',status_id
            print 'model  is ',id_
            print 'before try !!!!!!!'
            print 'currency is', currency
            try:
                p = Material.objects.get(id=id_)
                
                if item != p.order_item:
                  comment=comment+" Changed item from \""+p.order_item+"\" to\""+item+"\"."
                  p.order_item = item
                  update_list.append('order_item')

                if category != p.order_category:
                  comment=comment+" Changed category from \""+p.order_category+"\" to\""+category+"\"."
                  p.order_category = category
                  update_list.append('order_category')
                if sub_category != p.order_sub_category:
                  comment=comment+" Changed subcategory from \""+p.order_sub_category+"\" to\""+sub_category+"\"."
                  p.order_sub_category = sub_category
                  update_list.append('order_sub_category')
                if quantity != p.order_quantity:
                  comment=comment+" Changed quantity from \""+str(p.order_quantity)+"\" to\""+str(quantity)+"\"."
                  p.order_quantity = quantity
                  update_list.append('order_quantity')
                # if currency !=p.order_currency:
                #   p.order_currency = currency
                if price != p.order_unit_price:
                  comment=comment+" Changed price from \""+str(p.order_unit_price)+"\" to\""+str(price)+"\"."
                  p.order_unit_price = price
                  if p.order_unit_price == "":
                      p.order_unit_price =0.00
                  update_list.append('order_unit_price')
                if url != p.order_item_url:
                  comment=comment+" Changed URL from \""+str(p.order_item_url)+"\" to\""+str(url)+"\"."
                  p.order_item_url = url
                  update_list.append('order_item_url')
                if vendor != p.order_vendor:
                  comment=comment+" Changed vendor from \""+p.order_vendor+"\" to\""+vendor+"\"."
                  p.order_vendor = vendor
                  update_list.append('order_vendor')
                if est_lead_num and est_lead_days:
                  print "in here"
                  match = re.match(r"([0-9]+)([a-z]+)", p.est_lead_time , re.I)
                  print 'match is', match
                  if match:
                    items=match.groups()
                    print 'items is ',items
                    if est_lead_num!=items[0] or est_lead_days !=items[1]:
                      comment=comment+" Changed ELT from \""+p.est_lead_time+"\" to\""+str(est_lead_num)+str(est_lead_days)+"\"."
                      p.est_lead_time = str(est_lead_num)+str(est_lead_days)
                      update_list.append('est_lead_time')
                
                if int(priority) != p.order_priority.priority_id:
                  print type(priority), type(p.order_priority.priority_id)
                  o = OrderPriority.objects.get(priority_id=p.order_priority.priority_id)
                  n = OrderPriority.objects.get(priority_id=priority)
                  comment=comment+" Changed priority from \""+o.name+"\" to\""+n.name+"\"."
                  p.order_priority = n
                  update_list.append('order_priority')

                if int(status_id) != p.order_status.status_id:
                  print type(status_id), type(p.order_status.status_id)
                  o = OrderStatus.objects.get(status_id=p.order_status.status_id)
                  n = OrderStatus.objects.get(status_id=status_id)
                  comment=comment+" Changed priority from \""+o.name+"\" to\""+n.name+"\"."
                  p.order_status = n
                  update_list.append('order_status')
                # if p.order_currency == "":
                #   p.order_currency="Rs"
                
                if hsn != p.order_hsn:
                  comment=comment+" Changed HSN code from \""+p.order_hsn+"\" to\""+hsn+"\"."
                  p.order_hsn = hsn
                  update_list.append('order_hsn')

                p.save(update_fields=update_list)
                if comment !="":
                  print 'comments are not none'
                  author = request.user
                  c = MaterialComment(material=p,comment=comment,author=author)
                  c.save()
                  all_users = User.objects.all()
                  for i in all_users:
                    print i.username
                    t = ReadCommentTrack(order_id=p.id,project_id = project_data.id,user_name=i.username,read_flag="N")
                    t.save()
                order_message="Order Updated Successfully"
                
            except Material.DoesNotExist:
                print 'in exception blck'
                status = OrderStatus.objects.filter(status_id = int(status_id))
                p = Material(project_name=project_data,order_category=category,\
                          order_sub_category=sub_category,order_item=item,\
                          order_item_url=url,order_vendor = vendor,order_quantity=int(quantity),\
                          order_currency = currency,order_unit_price=float(price), order_status=status, \
                          est_lead_time=est_lead_time,order_priority=priority, order_hsn=hsn)
                p.save()
                order_message="Order Updated Successfully"
            print order_message
            request.session['order_message'] = order_message
            return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))
         else:
            print 'errror in herereeeee'
            print request.method
            order_message="Order Update Not Successful"
            print material_form.errors
            print 'something went wrong !!!!!!!!' 
            request.session['order_message'] = order_message  
            return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))
      if request.method=='POST' and 'order_delete' in request.POST:
          id_ = request.POST.get('model_instance')
          print'Name issss',id_
          material_data = Material.objects.get(id=id_)
          material_data.delete()
          order_message="Order Deleted Successfully"
          request.session['order_message'] = order_message
          return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))

      if request.method=='POST' and 'tracking_save' in request.POST:
          tracking_form = TrackingForm(request.POST)
          if tracking_form.is_valid():
            id_ = request.POST.get("model_instance")
            tracking_num = tracking_form.cleaned_data.get('tracking_number')
            slug_id = tracking_form.cleaned_data.get('slug')
            courier = Courier.objects.get(id=slug_id)
          print 'data is', courier.slug
          try:
             api = aftership.APIv4("8a4aab03-8132-4fae-aca8-48c054964e14")
             try:
               print api.trackings.post(tracking=dict(slug=courier.slug, tracking_number=tracking_num, title="Title"))
             except:
               print 'Tracking number already exists in API'
             print api.trackings.get(courier.slug, tracking_num)
             print api.trackings.get(courier.slug, tracking_num, fields=['tag','last_updated_at','location','updated_at','checkpoint_time','shipment_delivery_date','expected_delivery'])
             material_data = Material.objects.get(id=id_)
             track_data = TrackingInfo.objects.filter(material=material_data.id)
             print track_data
             
             # recording the tracking to database
             if track_data:
                 print "Tracking for material present"
                 track_data = TrackingInfo.objects.get(material=material_data.id)
                 track_data.tracking_no = tracking_num
                 track_data.courier_id = courier
                 track_data.save()
                 order_message="Tracking Updated Successfully"
 
             else:

                 p = TrackingInfo(tracking_no = tracking_num, courier_id = courier, material = material_data)
                 p.save()
                 order_message="Tracking Recorded Successfully"
             request.session['order_message'] = order_message
             return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))
          except Exception as e:
             print 'tracking for material doesnt exist'
             print tracking_form.errors
             print str(e) 
             order_message = "Tracking Not Found"
             request.session['order_message'] = order_message
             return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))
          #material_data = Material.objects.get(id=id_)
          #material_data.delete()
          
      if request.method=='POST' and 'prototype_save' in request.POST:
          prototype_form = PrototypeForm(request.POST, request.FILES)
          print 'method is post',request.POST
          print 'these are the files', request.FILES

          if prototype_form.is_valid():   
            print' this form is valid'
            pname = prototype_form.cleaned_data.get('pname')
            photo = request.FILES['photo']
            description = prototype_form.cleaned_data.get('description')
            print 'photo is ', photo
            p = Prototype(project_name=project_data,pname=pname,photo=photo,description=description)     
            print 'image is ', p.photo
            p.save()
            prototype_message="Photo Uploaded Successfully"      
          else:
            print 'errror in herereeeee'
            print request.method
            prototype_message="Photo could not be uploaded"
            print prototype_form.errors
          request.session['prototype_message'] = prototype_message
          return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))
      if request.method=='POST' and 'delete_photo' in request.POST:
          id_ = request.POST.get('model_instance')
          print'photo issss',id_
          prototype_data = Prototype.objects.get(id=id_)
          prototype_data.delete()
          prototype_message="Order Deleted Successfully"
          request.session['prototype_message'] = prototype_message
          return HttpResponseRedirect(reverse('projectdetails', args=[project_data.name ]))
          #message="Order Deleted Successfully"
      try:
         project_plan_data = None
         project_plan_data = Plan.objects.get(project_name=project_data.id)
      except Plan.DoesNotExist:
         project_plan_data = None   
      try:
         order_data = None
         print 'ffetching from hereeeeeee'
         order_data = Material.objects.filter(project_name=project_data.id)
      except Material.DoesNotExist:
         order_data = None
      try:
         schedule_data = None
         request.session['task_message'] = task_message
         schedule_data = Schedule.objects.filter(project_name=project_data.id)
      except Schedule.DoesNotExist:
         schedule_data = None
      try:   
         prototype_data = None
         prototype_data = Prototype.objects.filter(project_name=project_data.id)
      except Prototype.DoesNotExist:
         prototype_data = None

      try:   
         comment_data = None
         comment_data = ReadCommentTrack.objects.filter(project_id=project_data.id,user_name=request.user)
         print (comment_data)
      except ReadCommentTrack.DoesNotExist:
         comment_data = None
      
      plan_form = ProjectPlanForm(request.POST or None)
      schedule_form = ScheduleForm(request.POST or None)
      schedule_edit_form = ScheduleEditForm(request.POST or None)
      schedule_comment_form = ScheduleCommentForm(request.POST or None)
      material_comment_form = MaterialCommentForm(request.POST or None)
      material_form = MaterialForm(request.POST or None)
      prototype_form = PrototypeForm(request.POST or None)
      process_url_form = ProcessURLForm(request.POST or None)
      excel_form = ExcelForm(request.POST or None)
      tracking_form = TrackingForm(request.POST or None)
      context = {
         "project_data":project_data,
         "project_plan_data":project_plan_data,
         "plan_form":plan_form,
         "schedule_form":schedule_form,
         "schedule_edit_form":schedule_edit_form,
         "material_form":material_form,
         "schedule_data":schedule_data,
         "task_message":task_message,
         "plan_message":plan_message,
         "order_message":order_message,
         "order_data":order_data,
         "order_status_data":order_status_data,
         "prototype_form":prototype_form,
         "prototype_message":prototype_message,
         "prototype_data":prototype_data,
         "schedule_comment_form":schedule_comment_form,
         "material_comment_form":material_comment_form,
         "process_url_form": process_url_form,
         "excel_form":excel_form,
         "tracking_form":tracking_form,
         "status_message":status_message,
         "comment_data":comment_data,
      }
      return render(request,"project/details/base.html",context)
      