import sys
import os
import django
#path to the project folder
sys.path.append('C:/Users/acer-pc/Documents/ET/my_project_folder/cpm_django/src/')
#project name
os.environ['DJANGO_SETTINGS_MODULE'] = 'mysite.settings'
django.setup()

import openpyxl
#app name
from cpm.models import Courier
#from mysite import settings
#from django.core.management import setup_environ

#setup_environ(settings)

def excel_courier_loader(excel):
    wb = openpyxl.load_workbook(excel)
    sheets = wb.sheetnames
    print sheets
    sheet = wb.get_sheet_by_name('couriers (1)-1')
    for row in range(2, sheet.max_row + 1):
       slug = sheet['A'+str(row)].value
       name = sheet['B'+str(row)].value
       try:
          p = Courier(name=name,slug=slug)
          p.save() 
       except Exception as e:
        print str(e)


if __name__ == "__main__":
    #path to excel
    excel = "C:/Users/acer-pc/Documents/ET/my_project_folder/couriers.xlsx"
    excel_courier_loader(excel)